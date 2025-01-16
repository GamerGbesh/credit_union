from django.shortcuts import render, get_object_or_404, redirect, HttpResponse
from .models import *
from django.db.models import F, Sum, DecimalField, DateTimeField, ExpressionWrapper
from datetime import datetime
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.contrib import messages
from django.db.models.functions import Coalesce
from django.contrib.auth.decorators import login_required
from .forms import MemberForm
from django.db import transaction
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

# Create your views here.
def home(request):
    if request.user.is_authenticated:
        union = CreditUnionBalance.objects.filter(user_id=request.user).first()
        if union:
            total_contributions = Contribution.objects.filter(user_id=request.user).aggregate(Sum("amount"))["amount__sum"]
            total_loans = ApprovedLoan.objects.filter(user_id=request.user).aggregate(Sum("amount_left"))["amount_left__sum"]
            if not total_contributions:
                total_contributions = 0
            if not total_loans:
                total_loans = 0
            return render(request, "home.html", {"union": union, 
                                                "total_contribution": total_contributions, 
                                                "total_loan_debt": total_loans})
        else:
            return render(request, "home.html", {"current_year": 2024})
    else:
        return render(request, "home.html", {"current_year": 2024})


@login_required
@transaction.atomic
def add_member(request):
    """This function is used to add a new member to the database

    Args:
        request (HttpRequest): The request object

    Returns:
        HttpResponse: Returns the webpage that is rendered
    """
    if request.method == "POST":
        """
        Checks if a member already exists in the database, if not adds them
        Returns a message if the member already exists
        """
        form = MemberForm(request.POST)
        if form.is_valid():
            full_name = form.cleaned_data["full_name"]
            msisdn = form.cleaned_data["full_number"]
            email = form.cleaned_data["email"]
            dob = form.cleaned_data["dob"]
            if not msisdn:
                messages.error(request, "Please enter a phone number")
                return redirect("add_member")
            if Member.objects.filter(msisdn=msisdn).exists():
                messages.error(request, "Phone number already exists! Try with a different number")
                return redirect("add_member")
            
            if email:
                check_email = KYCDetails.objects.filter(email=email).exists()
            else:
                check_email = False
                email = None
            
            if check_email:
                messages.error(request, "Email already exists! Try with a different email")
                return redirect("add_member")
            user = request.user
            member = Member(msisdn=msisdn, user_id=user)
            details = KYCDetails(member_msisdn=member, name=full_name, email=email, dob=dob)
            details.save()
            member.save()
            
            
            return redirect("view_members")
    else:
        form = MemberForm()
    union = CreditUnionBalance.objects.filter(user_id=request.user).first()
    return render(request, "add_member.html", {"union": union, "form": form})



# All filters will be done using the enums set in models.py
@login_required
def view_members(request):
    # Create a filter for members
    sort = request.GET.get("sort", "name")
    order = request.GET.get("order", "")
    sorted = order + sort
    filter = request.GET.get("filter", MemberStatus.ACTIVE)
    
    members = Member.objects.filter(user_id=request.user).annotate(name=F("kycdetails__name"),
        total_contribution=Coalesce(Sum("contribution__amount", distinct=True), 0, output_field=DecimalField()),
        loan_debt=Coalesce(Sum("approvedloan__amount_left", distinct=True), 0, output_field=DecimalField())).order_by(sorted)
    
    if filter != "all":
        members = members.filter(status=filter).filter(user_id=request.user)
    
    paginator = Paginator(members, 10)
    page_number = request.GET.get("page")

    try:
        page_obj = paginator.get_page(page_number)
    except PageNotAnInteger:
        page_obj = paginator.page(1)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages)
  
    union = CreditUnionBalance.objects.filter(user_id=request.user).first()
    return render(request, "view_members.html", {"page_obj": page_obj, 
                                                 "sort":sort,
                                                 "order":order,
                                                 "filter":filter,
                                                 "union": union
                                                 })


@login_required
def member_info(request, msisdn):
    member = Member.objects.filter(user_id=request.user).annotate(name=F("kycdetails__name"),
                                                                    total_contribution=Coalesce(Sum("contribution__amount", distinct=True), 0, output_field=DecimalField()),
                                                                    loan_debt=Coalesce(Sum("approvedloan__amount_left", distinct=True), 0, output_field=DecimalField()),
                                                                    email=F("kycdetails__email"),
                                                                    dob=F("kycdetails__dob")
                                                                    ).get(msisdn=msisdn)
    union = CreditUnionBalance.objects.filter(user_id=request.user).first()
    return render(request, "member.html", {"member": member, "union": union})


@login_required
@transaction.atomic
def edit_member(request, msisdn):
    member = Member.objects.filter(user_id=request.user).annotate(name=F("kycdetails__name"),
                                     surname=F("kycdetails__surname"),
                                     other_names=F("kycdetails__other_names"),  
                                     email=F("kycdetails__email"), 
                                     dob=F("kycdetails__dob")
                                     ).get(msisdn=msisdn)
    
    if request.method == "POST":
        new_msisdn = (request.POST.get("phone"))
        email = request.POST.get("email")
        if not email:
            email = None
        surname = request.POST.get("surname")
        other_names = request.POST.get("other_names")
        member_obj = Member.objects.filter(user_id=request.user).get(msisdn=msisdn)
        member_obj.msisdn = new_msisdn
        member_obj.save()        
        kyc = get_object_or_404(KYCDetails, member_msisdn=member)
        kyc.email = email
        kyc.surname = surname
        kyc.other_names = other_names
        full_name = f"{surname} {other_names}"
        kyc.name = full_name
        kyc.member_msisdn = member_obj 
        kyc.save()        
        return redirect("members", msisdn=new_msisdn)
    union = CreditUnionBalance.objects.filter(user_id=request.user).first()
    return render(request, "edit_member.html", {"member": member, "union": union})


@login_required
@transaction.atomic
def withdraw(request, msisdn):
    member = Member.objects.filter(user_id=request.user).annotate(name=F("kycdetails__name"),
                                     balance=ExpressionWrapper(Coalesce(Sum("contribution__amount", distinct=True), 0), DecimalField()),
                                     loan_debt=ExpressionWrapper(Coalesce(Sum("approvedloan__amount_left", distinct=True), 0), DecimalField()),
                                     interest=ExpressionWrapper(Coalesce(Sum("contribution__interest", distinct=True), 0), DecimalField()),
                                     final_balance=ExpressionWrapper(Coalesce(Sum("contribution__amount", distinct=True), 0) + 
                                                   Coalesce(Sum("contribution__interest", distinct=True), 0) - 
                                                   Coalesce(Sum("approvedloan__amount_left", distinct=True), 0), DecimalField())
                                     ).get(msisdn=msisdn)
    if request.method == "POST":
        date = request.POST.get("date")
        if member.status == MemberStatus.INACTIVE:
            messages.error(request, f"{member.name} is not an active member")
            return redirect("home")
        if not date:
                date = datetime.now().date()

        credit_union_balance = CreditUnionBalance.objects.filter(user_id=request.user).first()
        if float(credit_union_balance.amount) < float(member.final_balance):
            messages.error(request, "The credit union balance is less than the total savings of the member cannot withdraw")
            messages.info(request, f"Credit union balance: {credit_union_balance.amount} Member savings: {member.final_balance}")
            return redirect("withdraw", msisdn=msisdn)
        if member.final_balance < 0:
            messages.error(request, "The member has a negative balance and cannot withdraw")
            return redirect("withdraw", msisdn=msisdn)
        credit_union_balance.amount = float(credit_union_balance.amount) - float(member.final_balance)
        
        credit_union_balance.save()
        contributions = Contribution.objects.filter(member_msisdn=member)
        for contribution in contributions:
                contribution.amount = 0
                contribution.interest = 0
                contribution.save()
        if member.loan_debt > 0:           
            loans = ApprovedLoan.objects.filter(member_msisdn=member)
            for loan in loans:
                loan.status = StatusEnum.PAID
                loan.amount_left = 0
                loan.save()
            Transaction.objects.create(member_msisdn=member, date=date, 
                                   transaction_type=TransactionEnum.LOAN_PAYMENT, 
                                   amount=member.loan_debt, 
                                   description="Payment of all loan debt")        
        Transaction.objects.create(member_msisdn=member, date=date, 
                                   transaction_type=TransactionEnum.SAVINGS_WITHDRAWAL, 
                                   amount=member.final_balance, 
                                   description="Withdrawal of total savings")
        member.status = MemberStatus.INACTIVE
        member.save()
        return redirect("members", msisdn)
    union = CreditUnionBalance.objects.filter(user_id=request.user).first()
    return render(request, "withdraw.html", {"member": member, "union": union})

# Contributions
@login_required
@transaction.atomic
def make_contribution(request):
    # add ability to change the date when making a contribution
    members = Member.objects.filter(user_id=request.user).annotate(name=F("kycdetails__name"))
    if request.method == "POST":
        name = request.POST.get("member")
        amount = float(request.POST.get("amount"))
        date = request.POST.get("date")
        interest = float(request.POST.get("interest"))/100
        interest = interest * amount
        if amount < 0:
            messages.error(request, "Amount cannot be negative")
            return redirect("make_contribution")
        
        if not date:
                date = str(datetime.now().date())

        try:
            member = Member.objects.filter(user_id=request.user).get(kycdetails__name=name)
        except Member.DoesNotExist:
            messages.error(request, "Member not found")
            return redirect("make_contribution")
        except Member.MultipleObjectsReturned:
            messages.info(request, "Multiple members have that name, select which of them!")
            members = Member.objects.filter(user_id=request.user, kycdetails__name=name).annotate(
                name=F("kycdetails__name"),
                balance=ExpressionWrapper(Coalesce(Sum("contribution__amount", distinct=True), 0), DecimalField()),
                loan_debt=ExpressionWrapper(Coalesce(Sum("approvedloan__amount_left", distinct=True), 0), DecimalField()),
            )
            paginator = Paginator(members, 10)
            page_number = request.GET.get("page")

            try:
                page_obj = paginator.get_page(page_number)
            except PageNotAnInteger:
                page_obj = paginator.page(1)
            except EmptyPage:
                page_obj = paginator.page(paginator.num_pages)
            union = CreditUnionBalance.objects.filter(user_id=request.user).first()
            return render(request, "choice.html", {"page_obj": page_obj, "union": union, "amount":amount, "interest":interest, "date":date})
        
        Contribution.objects.create(member_msisdn=member, amount=amount, interest=interest, user_id=request.user)
        Transaction.objects.create(member_msisdn=member, 
                                       transaction_type=TransactionEnum.DEPOSIT, 
                                       amount=amount, 
                                       description="Contribution",
                                       date=date)
        union = CreditUnionBalance.objects.filter(user_id=request.user).first()
        union.amount = float(union.amount) + amount
        union.save()
        messages.success(request, "Contribution made successfully")
        return redirect("view_members")
    union = CreditUnionBalance.objects.filter(user_id=request.user).first()
    return render(request, "make_contribution.html", {"members": members, "union": union})


def choice(request):
    if request.method == "POST":
        choice = request.POST.get("choice")
        member = Member.objects.filter(user_id=request.user).get(msisdn=choice)
        amount = float(request.POST.get("amount"))
        interest = float(request.POST.get("interest"))
        date = request.POST.get("date")
        
        Contribution.objects.create(member_msisdn=member, amount=amount, interest=interest, user_id=request.user)
        Transaction.objects.create(member_msisdn=member, 
                                       transaction_type=TransactionEnum.DEPOSIT, 
                                       amount=amount, 
                                       description="Contribution",
                                       date=date)
        union = CreditUnionBalance.objects.filter(user_id=request.user).first()
        union.amount = float(union.amount) + amount
        union.save()
        messages.success(request, "Contribution made successfully")
        return redirect("view_members")


@login_required
def delete(request, msisdn):
    member = Member.objects.filter(user_id=request.user).annotate(name=F("kycdetails__name"),
                                     balance=ExpressionWrapper(Coalesce(Sum("contribution__amount", distinct=True), 0), DecimalField()),
                                     loan_debt=ExpressionWrapper(Coalesce(Sum("approvedloan__amount_left", distinct=True), 0), DecimalField()),
                                     interest=ExpressionWrapper(Coalesce(Sum("contribution__interest", distinct=True), 0), DecimalField()),
                                     final_balance=ExpressionWrapper(Coalesce(Sum("contribution__amount", distinct=True), 0) + 
                                                   Coalesce(Sum("contribution__interest", distinct=True), 0) - 
                                                   Coalesce(Sum("approvedloan__amount_left", distinct=True), 0), DecimalField())
                                     ).get(msisdn=msisdn)
    if request.method == "POST":
        if member.balance == 0 and member.loan_debt == 0 and member.final_balance == 0:
            member.delete()
            messages.success(request, f"{member.name} deleted successfully!")
            return redirect("home")
        else:
            messages.error(request, "Member has active contributions or loans to pay! Member cannot be deleted!")
            messages.info(request, "To delete first withdraw money for the member!")
            return redirect("members", msisdn)
    union = CreditUnionBalance.objects.filter(user_id=request.user).first()
    return render(request, "delete.html", {"member": member, "union": union})





# History
@login_required
def view_history(request):
    #Create a filter and sort the transactions
    sort = request.GET.get("sort", "transaction_date")
    order = request.GET.get("order", "-")
    sorted = order + sort
    filter = request.GET.get("filter", "all")
    members = Member.objects.filter(user_id=request.user).annotate(name=F("kycdetails__name"), 
                                      transaction_type=F("transaction__transaction_type"),  
                                      description=F("transaction__description"), 
                                      amount=F("transaction__amount"),
                                      date=F("transaction__date"),
                                      transaction_date = ExpressionWrapper(F("transaction__date") + F("transaction__time"), DateTimeField())).filter(amount__gt=0).order_by(sorted)

    if filter != "all":
        members = members.filter(transaction_type=filter)
    paginator = Paginator(members, 20)
    page_number = request.GET.get("page")

    try:
        page_obj = paginator.get_page(page_number)
    except PageNotAnInteger:
        page_obj = paginator.page(1)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages)
    union = CreditUnionBalance.objects.filter(user_id=request.user).first()
    return render(request, "history.html", {"page_obj": page_obj,
                                            "sort":sort,
                                            "order":order,
                                            "filter":filter,
                                            "union": union})



# Loans
@login_required
def loan_request(request):
    sort = request.GET.get("sort", "loan_created")
    order = request.GET.get("order", "-")
    sorted = order + sort
    filter = request.GET.get("filter", "all")
    members = Member.objects.filter(user_id=request.user).annotate(name=F("kycdetails__name"), 
                                      amount_requested=F("loanrequest__amount_requested"), 
                                      loan_status=F("loanrequest__status"), 
                                      date=F("loanrequest__request_date"), 
                                      loan_id=F("loanrequest__id"),
                                      loan_created=F("loanrequest__created")).filter(amount_requested__gt=0).order_by(sorted)
    if filter != "all":
        members = members.filter(loan_status=filter)

    paginator = Paginator(members, 10)
    page_number = request.GET.get("page")

    try:
        page_obj = paginator.get_page(page_number)
    except PageNotAnInteger:
        page_obj = paginator.page(1)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages)
    union = CreditUnionBalance.objects.filter(user_id=request.user).first()
    return render(request, "loans_requests.html", {"page_obj": page_obj,
                                          "sort":sort,
                                          "order":order,
                                          "filter":filter,
                                          "union":union})


@login_required
def approved_loans(request):
    sort = request.GET.get("sort", "updated")
    order = request.GET.get("order", "-")
    sorted = order + sort
    filter = request.GET.get("filter", "all")
    members = ApprovedLoan.objects.annotate(name=F("member_msisdn__kycdetails__name"), 
                                      total_amount=F("amount_of_loan") + F("interest"),
                                      loan_id=F("loan_request_id"),
                                      ).filter(amount_of_loan__gt=0).order_by(sorted).filter(user_id=request.user)
    if filter != "all":
        members = members.filter(status=filter)

    paginator = Paginator(members, 10)
    page_number = request.GET.get("page")

    try:
        page_obj = paginator.get_page(page_number)
    except PageNotAnInteger:
        page_obj = paginator.page(1)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages)
    union = CreditUnionBalance.objects.filter(user_id=request.user).first()
    return render(request, "loans.html", {"page_obj": page_obj,
                                          "sort":sort,
                                          "order":order,
                                          "filter":filter,
                                          "union":union})

@login_required
@transaction.atomic
def loan_details(request, loan_id):
    loan = LoanRequest.objects.annotate(name=F("member_msisdn__kycdetails__name"), 
                                        loan_description=F("loan_purpose"),
                                        loan_status=F("status"), 
                                        date=F("request_date"), 
                                        loan_id=F("id")).get(id=loan_id)
    member = Member.objects.filter(user_id=request.user).get(msisdn=loan.member_msisdn.msisdn)
    year = datetime.now().year + 1
    month = datetime.now().month
    day = datetime.now().day
    due_date = datetime(year, month, day).date()
    try:
        approved = ApprovedLoan.objects.get(loan_request_id=loan)
        if approved.amount_left == 0 and approved.status != StatusEnum.PAID:
            approved.status = StatusEnum.PAID  
            approved.save()
        elif datetime.now().date() > approved.end_of_loan_date and approved.status != StatusEnum.PAID:
            approved.status = "OVERDUE"
            approved.save()
        if approved.status == StatusEnum.PAID:
            messages.success(request, "Loan has been paid")
        elif approved.status == "OVERDUE":
            messages.error(request, "Loan is overdue")

    except ApprovedLoan.DoesNotExist:
        approved = None
    
    
    if request.method == "POST":
        value = request.POST.get("button")
        if value == "approve":
            credit_union_balance = CreditUnionBalance.objects.filter(user_id=request.user).first()
            if credit_union_balance.amount <= loan.amount_requested:
                messages.error(request, "The amount being requested for loan is more than the available balance in the credit union!")
                messages.info(request, f"Amount available in the credit union: {credit_union_balance.amount} Amount being requested: {loan.amount_requested}")
                return redirect("loan_details", loan_id=loan_id)
            loan.status = LoanStatus.APPROVED
            loan.save()
            date = request.POST.get("date")
            interest_rate = float(request.POST.get("interest"))/100
            if interest_rate < 0:
                messages.error(request, "Interest rate cannot be negative")
                return redirect("loan_details", loan_id=loan_id)
            date_of_loan = request.POST.get("due_date")
            if date_of_loan:
                due_date = date_of_loan

            
            member = Member.objects.filter(user_id=request.user).get(msisdn=loan.member_msisdn.msisdn)
            if not date:
                date = datetime.now().date()
            
            amount_to_deduct, interest, months, date_accepted = prorating(loan.amount_requested, interest_rate, due_date, date)
            

            ApprovedLoan.objects.create(member_msisdn=member,
                                        loan_request_id=loan, 
                                        amount_of_loan=loan.amount_requested, 
                                        interest=interest, 
                                        end_of_loan_date=due_date,
                                        monthly_deduction=amount_to_deduct/months,
                                        amount_left=amount_to_deduct,
                                        user_id=request.user,
                                        created=date_accepted)
            credit_union_balance.amount = float(credit_union_balance.amount) - float(loan.amount_requested)
            credit_union_balance.save()
            Transaction.objects.create(member_msisdn=member, 
                                       transaction_type=TransactionEnum.LOAN_WITHDRAWAL, 
                                       amount=loan.amount_requested, 
                                       description=loan.loan_description,
                                       date=date)
            approved = ApprovedLoan.objects.filter(loan_request_id=loan)
            
            
        elif value == "reject":
            loan.status = LoanStatus.REJECTED
            loan.save()
        return redirect("loan_details", loan_id=loan_id)
    union = CreditUnionBalance.objects.filter(user_id=request.user).first()
    return render(request, "loan_details.html", {"loan": loan, "approved": approved, "due_date": due_date, "union": union})

@login_required
@transaction.atomic
def input_loans(request):
   
    members = Member.objects.filter(user_id=request.user).annotate(name=F("kycdetails__name"))
    if request.method == "POST":
        name = request.POST.get("member")
        amount = float(request.POST.get("loan_amount"))
        date = request.POST.get("date")
        if not date:
                date = str( datetime.now().date())
        loan_purpose = request.POST.get("loan_purpose")
        if amount < 0:
            messages.error(request, "Loan amount cannot be negative")
            return redirect("input_loans")
        try:
            member = Member.objects.filter(user_id=request.user, status=MemberStatus.ACTIVE).get(kycdetails__name=name)
        except Member.DoesNotExist:
            messages.error(request, "Member not found")
            return redirect("input_loans")
        except Member.MultipleObjectsReturned:
            messages.info(request, "Multiple members have that name, select which of them!")
            members = Member.objects.filter(user_id=request.user, kycdetails__name=name).annotate(
                name=F("kycdetails__name"),
                balance=ExpressionWrapper(Coalesce(Sum("contribution__amount", distinct=True), 0), DecimalField()),
                loan_debt=ExpressionWrapper(Coalesce(Sum("approvedloan__amount_left", distinct=True), 0), DecimalField()),
            )
            paginator = Paginator(members, 10)
            page_number = request.GET.get("page")

            try:
                page_obj = paginator.get_page(page_number)
            except PageNotAnInteger:
                page_obj = paginator.page(1)
            except EmptyPage:
                page_obj = paginator.page(paginator.num_pages)
            union = CreditUnionBalance.objects.filter(user_id=request.user).first()
            return render(request, "loan_choice.html", {"page_obj": page_obj, "union": union, "amount":amount, "loan_amount":amount, "loan_purpose":loan_purpose, "date":date})        

        LoanRequest.objects.create(member_msisdn=member, 
                                   amount_requested=amount, 
                                   loan_purpose=loan_purpose, 
                                   request_date=date, 
                                   )

        return redirect("loan_request")
    union = CreditUnionBalance.objects.filter(user_id=request.user).first()
    return render(request, "input_loans.html", {"members": members, "union": union})

@login_required
def loan_choice(request):
    if request.method == "POST":
        choice = request.POST.get("choice")
        member = Member.objects.filter(user_id=request.user).get(msisdn=choice)
        amount = float(request.POST.get("loan_amount"))
        loan_purpose = request.POST.get("loan_purpose")
        date = request.POST.get("date")
        
        LoanRequest.objects.create(member_msisdn=member, 
                                   amount_requested=amount, 
                                   loan_purpose=loan_purpose, 
                                   request_date=date, 
                                   )
        return redirect("loan_request")


@login_required
@transaction.atomic
def pay_loan(request, loan_id):
    loan = LoanRequest.objects.get(id=loan_id)
    approved = ApprovedLoan.objects.get(loan_request_id=loan)
    # If there is payment to be made, this POST checks the validity of the payment and ensures it's paid
    if request.method == "POST":
        amount_paid = float(request.POST.get("amount"))
        date = request.POST.get("date")
        # Making sure the payment value is valid
        if amount_paid < 0:
            messages.error(request, "Amount paid cannot be negative")
            return redirect("pay_loan", loan_id=loan_id)
        approved = ApprovedLoan.objects.get(loan_request_id=loan_id)
        if amount_paid > approved.amount_left:
            messages.error(request, "The amount being paid is greater than what is left of the loan to be paid")
            messages.info(request, "Amount of loan left to be paid: GHc" + str(approved.amount_left) + ", Amount being paid: GHc" + str(amount_paid))

            return redirect("pay_loan", loan_id=loan_id)
        # Making sure all transactions are reflecting
        approved.amount_left = float(approved.amount_left) - amount_paid
        approved.save()
        if not date:
                date = datetime.now().date()
        Transaction.objects.create(member_msisdn=loan.member_msisdn, 
                                       transaction_type=TransactionEnum.LOAN_PAYMENT, 
                                       amount=amount_paid, description="Loan payment", 
                                       date=date)
        credit_union_balance = CreditUnionBalance.objects.filter(user_id=request.user).first()
        credit_union_balance.amount = float(credit_union_balance.amount) + amount_paid
        credit_union_balance.save()
        return redirect("loan_details", loan_id=loan_id)
    union = CreditUnionBalance.objects.filter(user_id=request.user).first()
    return render(request, "pay_loan.html", {"loan": loan, "union": union})



def prorating(amount, interest_rate, due_date, date_accepted):
    """Used to calculate the amount to be paid over the duration of the loan taken

    Args:
        amount (float): Amount of loan taken
        interest_rate (float): Interest rate on the loan
        due_date (datetime.date | str): The date the loan is due
        date_accepted (datetime.date | str): The date the loan was accepted

    Returns:
        The total amount, interest, number of months for payment and date accepted
    """
    amount = float(amount)
    if isinstance(due_date, str):
        due_year, due_month, due_day = due_date.split("-")
        due_date = datetime(int(due_year), int(due_month), int(due_day)).date()
    if isinstance(date_accepted, str):  
        date_year, date_month, date_day = date_accepted.split("-")
        date_accepted = datetime(int(date_year), int(date_month), int(date_day)).date()
    interest = amount * interest_rate
    months = round((due_date - date_accepted).days / 30)
    interest = interest * (months / 12)
    total = amount + interest
    return total, interest, months, date_accepted


def contributions_to_excel(request):
    
    """This function converts the contributions of members into an excel sheet

    Args:
        request (HttpResponse): 

    Returns:
        HttpResponse: Downloadable excel sheet
    """
    members = Member.objects.filter(user_id=request.user).annotate(name=F("kycdetails__name"),
                                                                    total_contribution=Coalesce(Sum("contribution__amount", distinct=True), 0, output_field=DecimalField()),
                                                                    loan_debt=Coalesce(Sum("approvedloan__amount_left", distinct=True), 0, output_field=DecimalField()),
                                                                    email=F("kycdetails__email"),
                                                                    dob=F("kycdetails__dob")
                                                                    ) 
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.cell(row=1, column=1).value = "Name"
    ws.cell(row=1, column=2).value = "Phone Number"
    ws.cell(row=1, column=3).value = "Total Contribution"
    ws.cell(row=1, column=4).value = "Loan Debt"
    ws.cell(row=1, column=5).value = "Email"
    ws.cell(row=1, column=6).value = "Date of Birth"
    ws.cell(row=1, column=7).value = "Status"
    
    for i in range(1, 9):
        ws.cell(row=1, column=i).font = Font(b=True)
        ws.column_dimensions[get_column_letter(i)].width += 5

    for i, member in enumerate(members, 2):
        ws.cell(row=i, column=1).value = member.name
        ws.cell(row=i, column=2).value = member.msisdn
        ws.cell(row=i, column=3).value = member.total_contribution
        ws.cell(row=i, column=4).value = member.loan_debt
        ws.cell(row=i, column=5).value = member.email
        ws.cell(row=i, column=6).value = member.dob
        ws.cell(row=i, column=7).value = member.status
        
        
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="contribution.xlsx"'
    
    # Save the workbook to the response
    wb.save(response)
    
    return response
    

def requests_to_excel(request):
    
    """This function converts the loan requests into an excel sheet

    Args:
        request (HttpResponse): 

    Returns:
        HttpResponse: Downloadable excel sheet
    """
    members = Member.objects.filter(user_id=request.user).annotate(name=F("kycdetails__name"), 
                                      amount_requested=F("loanrequest__amount_requested"), 
                                      loan_status=F("loanrequest__status"), 
                                      date=F("loanrequest__request_date"), 
                                      loan_id=F("loanrequest__id"),
                                      loan_created=F("loanrequest__created")).filter(amount_requested__gt=0)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.cell(row=1, column=1).value = "Name"
    ws.cell(row=1, column=2).value = "Amount Requested"
    ws.cell(row=1, column=3).value = "Loan Status"
    ws.cell(row=1, column=4).value = "Date"

    for i in range(1, 6):
        ws.cell(row=1, column=i).font = Font(b=True)
        ws.column_dimensions[get_column_letter(i)].width += 5

    for i, member in enumerate(members, 2):
        ws.cell(row=i, column=1).value = member.name
        ws.cell(row=i, column=2).value = member.amount_requested
        ws.cell(row=i, column=3).value = member.loan_status
        ws.cell(row=i, column=4).value = member.date
   
        
        
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="loan Requests.xlsx"'
    
    # Save the workbook to the response
    wb.save(response)
    
    return response

def loans_to_excel(request):
    
    """This function converts the loans into an excel sheet

    Args:
        request (HttpResponse): 

    Returns:
        HttpResponse: Downloadable excel sheet
    """
    members = ApprovedLoan.objects.annotate(name=F("member_msisdn__kycdetails__name"), 
                                      total_amount=F("amount_of_loan") + F("interest"),
                                      loan_id=F("loan_request_id"),
                                      ).filter(amount_of_loan__gt=0).filter(user_id=request.user)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.cell(row=1, column=1).value = "Name"
    ws.cell(row=1, column=2).value = "Amount of Loan"
    ws.cell(row=1, column=3).value = "Interest"
    ws.cell(row=1, column=4).value = "Total Amount"
    ws.cell(row=1, column=5).value = "Amount Left"
    ws.cell(row=1, column=6).value = "Status"
    
    for i in range(1, 8):
        ws.cell(row=1, column=i).font = Font(b=True)
        ws.column_dimensions[get_column_letter(i)].width += 5

    for i, member in enumerate(members, 2):
        ws.cell(row=i, column=1).value = member.name
        ws.cell(row=i, column=2).value = member.amount_of_loan
        ws.cell(row=i, column=3).value = member.interest
        ws.cell(row=i, column=4).value = member.total_amount
        ws.cell(row=i, column=5).value = member.amount_left
        ws.cell(row=i, column=6).value = member.status        
        
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="loans.xlsx"'
    
    # Save the workbook to the response
    wb.save(response)
    
    return response

def history_to_excel(request):
    """This function converts the history of transactions into an excel sheet

    Args:
        request (HttpResponse): 

    Returns:
        HttpResponse: Downloadable excel sheet
    """
    members = Member.objects.filter(user_id=request.user).annotate(name=F("kycdetails__name"), 
                                      transaction_type=F("transaction__transaction_type"),  
                                      description=F("transaction__description"), 
                                      amount=F("transaction__amount"),
                                      date = F("transaction__date")).filter(amount__gt=0)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.cell(row=1, column=1).value = "Name"
    ws.cell(row=1, column=2).value = "Transaction Type"
    ws.cell(row=1, column=3).value = "Description"
    ws.cell(row=1, column=4).value = "Amount"
    ws.cell(row=1, column=5).value = "Date"
    
    for i in range(1, 7):
        ws.cell(row=1, column=i).font = Font(b=True)
        ws.column_dimensions[get_column_letter(i)].width += 5

    for i, member in enumerate(members, 2):
        ws.cell(row=i, column=1).value = member.name
        ws.cell(row=i, column=2).value = member.transaction_type
        ws.cell(row=i, column=3).value = member.description
        ws.cell(row=i, column=4).value = member.amount
        ws.cell(row=i, column=5).value = member.date

        
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="history.xlsx"'
    
    # Save the workbook to the response
    wb.save(response)
    
    return response


def integrate(request):
    info = KYCDetails.objects.all()
    for member in info:
        first_name = member.surname
        last_name = member.other_names
        full_name = f"{first_name} {last_name}"
        member.name = full_name
        member.save()

    return HttpResponse("<h1>Done</h1>")